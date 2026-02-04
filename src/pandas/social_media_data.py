"""
社交媒体数据处理
"""
from dataclasses import dataclass
from pathlib import Path
from typing import Optional
import os
from openpyxl import load_workbook
import pandas as pd

from tools.http_tool import query_match_info

p = Path(__file__).resolve()
current_dir = p.parents[2]


# 排除的关键词
EXCLUDED_KEY_WORD = ["日报", "其他", "节目", "抽签仪式"]

@dataclass
class ExcelModel():
    """
    表示excel文件中的一个工作表

    Attributes:
        sheet_index (int): 工作表索引
        sheet_name (str): 工作表名称
        df (pd.DataFrame): 工作表数据
        key (str): 唯一标识
    """
    sheet_index: int
    sheet_name: str
    df: pd.DataFrame
    key: Optional[str] = None


def read_social_midia_excel(file_path: str, header_row: int = 0) -> list[ExcelModel]:
    """
    读取excel文件，通过表头关键字锁定社媒数据sheet，并读取数据。
    该方式会将数字字符串自动转换成小数。并移除整行为空的行

    Args:
        file_path (str): excel文件路径
        header_row (int, optional): 表头所在行号，默认第0行. Defaults to 0.

    Returns:
        list[ExcelModel]: 包含所有匹配sheet的列表，每个元素为一个ExcelModel对象
    """
    xls = pd.ExcelFile(file_path)
    matched_sheets = []
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            nrows=header_row  # 只读表头，极快
        )

        cols = list(map(str, df.columns))
        # 抖音直播数据规则：B列为 “直播名称”、J列为 “直播方式”，R列为 “互动数据”
        if "抖音" in sheet_name and "直播名称" in cols and "直播方式" in cols and "互动数据" in cols:
            print(f"sheet_name: {sheet_name}, col size: {len(cols)}")
            if len(cols) != 36:
                raise ValueError(f"sheet_name: {sheet_name}, col size: {len(cols)}, 不符合抖音直播数据格式要求，请检查是否缺少或多余列。")
            else:
              full_df = pd.read_excel(
                  xls,
                  sheet_name=sheet_name,
                  header=[0, 1],
                  dtype=str
              ).dropna(how='all').reset_index(drop=True) # 移除整行为空的行

              # 去掉字符串列中的tab符
              str_cols = full_df.select_dtypes(include="string").columns
              full_df[str_cols] = full_df[str_cols].apply(
                  lambda s: s.str.replace("\t", " ", regex=False)
              )

              matched_sheets.append(ExcelModel(
                  sheet_index=100+xls.sheet_names.index(sheet_name), # index同一加上100，便于后续处理id
                  sheet_name=sheet_name,
                  df=full_df,
                  key="douyin"
              ))

        # 微信视频号直播数据规则：B列为 “直播名称”、J列为 “直播方式”，R列为 “互动数据”
        if "视频号" in sheet_name and "直播名称" in cols and "直播方式" in cols and "互动数据" in cols:
            print(f"sheet_name: {sheet_name}, col size: {len(cols)}")
            if len(cols) != 27:
                raise ValueError(f"sheet_name: {sheet_name}, col size: {len(cols)}, 不符合微信视频号直播数据格式要求，请检查是否缺少或多余列。")
            else:
              full_df = pd.read_excel(
                  xls,
                  sheet_name=sheet_name,
                  header=[0, 1],
                  dtype=str
              ).dropna(how='all').reset_index(drop=True) # 移除整行为空的行

              # 去掉字符串列中的tab符
              str_cols = full_df.select_dtypes(include="string").columns
              full_df[str_cols] = full_df[str_cols].apply(
                  lambda s: s.str.replace("\t", " ", regex=False)
              )
              matched_sheets.append(ExcelModel(
                  sheet_index=100+xls.sheet_names.index(sheet_name), # index同一加上100，便于后续处理id
                  sheet_name=sheet_name,
                  df=full_df,
                  key="wechat"
              ))

        # 小红书直播数据规则：B列为 “直播名称”、J列为 “直播方式”，T列为 “互动数据”
        if "小红书" in sheet_name and "直播名称" in cols and "直播方式" in cols and "互动数据" in cols:
            print(f"sheet_name: {sheet_name}, col size: {len(cols)}")
            if len(cols) != 38:
                raise ValueError(f"sheet_name: {sheet_name}, col size: {len(cols)}, 不符合小红书直播数据格式要求，请检查是否缺少或多余列。")
            else:
              full_df = pd.read_excel(
                  xls,
                  sheet_name=sheet_name,
                  header=[0, 1],
                  dtype=str
              ).dropna(how='all').reset_index(drop=True) # 移除整行为空的行

              # 去掉字符串列中的tab符
              str_cols = full_df.select_dtypes(include="string").columns
              full_df[str_cols] = full_df[str_cols].apply(
                  lambda s: s.str.replace("\t", " ", regex=False)
              )
              matched_sheets.append(ExcelModel(
                  sheet_index=100+xls.sheet_names.index(sheet_name), # index同一加上100，便于后续处理id
                  sheet_name=sheet_name,
                  df=full_df,
                  key="xiaohongshu"
              ))

    if not matched_sheets:
        raise ValueError("未找到包含指定表头的sheet")

    return matched_sheets


def read_social_midia_excel_with_openpyxl(file_path: str) -> list[ExcelModel]:
    """
    读取excel文件，通过表头关键字锁定社媒数据sheet，并读取数据。
    该方式会将保留数字原始格式，不进行自动转换（计算公式的除外）。

    Args:
        file_path (str): excel文件路径

    Returns:
        list[ExcelModel]: 包含所有匹配sheet的列表，每个元素为一个ExcelModel对象
    """
    matched_sheets = []
    col_size = 36

    file_path = os.path.expanduser(file_path)
    wb_formula = load_workbook(file_path, data_only=False)
    wb_value   = load_workbook(file_path, data_only=True)

    rows = []
    for sheet_name in wb_formula.sheetnames:
        ws_f = wb_formula[sheet_name]
        # 提取表头文本
        header_text = []
        row_cells = ws_f[1]  # openpyxl行号从1开始，header_rows是0索引
        cols = [cell.value for cell in row_cells]
        # 抖音直播数据规则：B列为 “直播名称”、J列为 “直播方式”，R列为 “互动数据”
        if "抖音" in sheet_name and "直播名称" in cols and "直播方式" in cols and "互动数据" in cols:
            print(f"====sheet_name: {sheet_name}, col size: {len(cols)}")
            if cols[col_size-1] != "备注":
                raise ValueError(f"sheet_name: {sheet_name}, col size: {len(cols)}, 不符合抖音直播数据格式要求，请检查是否缺少或多余列。")
            else:
                ws_v = wb_value[sheet_name]
                # 只取前col_size列数据
                for r_f, r_v in zip(ws_f.iter_rows(max_col=col_size), ws_v.iter_rows(max_col=col_size)):
                    row = []
                    for c_f, c_v in zip(r_f, r_v):
                        if c_f.data_type == "f":
                            row.append(c_v.value)   # 公式 → 结果（数值）
                        else:
                            row.append("" if c_f.value is None else str(c_f.value))

                    # 判断整行是否为空
                    if any(row):
                        rows.append(row)
                        #print(row)


def test_process_data(item: ExcelModel):
    """
    数据处理联系

    Args:
        item (ExcelModel): 包含数据的ExcelModel对象

    Returns:
        print
    """

    print(f"sheet_name: {item.sheet_name}")
    print("\n")
    print("="*50)

    print(item.df.columns)
    print("\n")
    print("="*50)

    for col1, col2 in item.df.columns:
        field = f"{col1}"
        if not col2.startswith("Unnamed"):
            field = f"{field}_{col2}"
        print(field)

    print("\n")
    print("="*50)
    df_audience = item.df["观众数据"]
    print(df_audience.head())

    #print(df.values)

    print("\n\n")


def process_douyin_data(item: ExcelModel):
    """
    处理抖音数据，将excel中数据转成hive数仓中的观众数据表
    1. 不要表头
    2. 每列之间用制表符隔开
    3. NaN值或其它特殊符号用空字符串替换
    4. 将处理后的数据写入文本文件

    Args:
        item (ExcelModel): 包含抖音数据的ExcelModel对象
    """
    df = item.df
    # 第一列为合并列时，统一处理成上一行的值
    df.iloc[:, 0] = df.iloc[:, 0].ffill()

    # 第一列中有换行符的，根据换行符分割，并取第一个元素作为第一列值
    df.iloc[:, 0] = df.iloc[:, 0].apply(lambda x: x.split('\n')[0] if isinstance(x, str) else x)

    # 其它 NaN 统一转为空字符串
    #df_audience = item.df.fillna("")

    for col in df.columns:
        l0, l1 = col
        # NaN 值统一转为空字符串
        df[col] = df[col].fillna("")
        # "/" 替换为空字符串
        df[col] = df[col].replace("/", "", regex=False)

        # 数值单位统一，将“万”单位转换成“个”整数
        if "（万）" in l0:
            df[col] = pd.to_numeric(df[col], errors="coerce").mul(10000).round(0).astype("Int64").astype(str).replace("<NA>", "")

    # 过滤第2列或第3列为空的行
    mask = (df.iloc[:, 1] == "") | (df.iloc[:, 2] == "")
    if mask.any():
        print("删除第2列或第3列为空的行：")
        #print(df[mask])
        df = df[~mask]

    # 处理列值为 "/" 的情况
    #obj_cols = df_audience.select_dtypes(include="object").columns
    #df_audience[obj_cols] = df_audience[obj_cols].replace("/", "", regex=False)

    # 新增id字段，并输出到第一列。id生成规则：sheet_index_ + (index + 1000)
    df2 = df.copy()
    # 插入到第0列
    df2.insert(0, "id", f"{item.sheet_index}_" + (df2.index + 1000).astype(str))

    df2.to_csv(
        f"{current_dir}/assets/data/social_midia/{item.key}_{item.sheet_index}.txt",
        sep="\t",
        index=False,
        header=False,
        encoding="utf-8"
    )


def process_wechat_data(item: ExcelModel):
    """
        处理视频号数据，将excel中数据转成hive数仓中的观众数据表
        1. 不要表头
        2. 每列之间用制表符隔开
        3. NaN值或其它特殊符号用空字符串替换
        4. 将处理后的数据写入文本文件

        Args:
            item (ExcelModel): 包含抖音数据的ExcelModel对象
    """
    df = item.df
    # 第一列为合并列时，统一处理成上一行的值
    df.iloc[:, 0] = df.iloc[:, 0].ffill()

    # 第一列中有换行符的，根据换行符分割，并取第一个元素作为第一列值
    df.iloc[:, 0] = df.iloc[:, 0].apply(lambda x: x.split('\n')[0] if isinstance(x, str) else x)

    for col in df.columns:
        l0, l1 = col
        # NaN 值统一转为空字符串
        df[col] = df[col].fillna("")
        # "/" 替换为空字符串
        df[col] = df[col].replace("/", "", regex=False)

        # 数值单位统一，将“万”单位转换成“个”整数
        if "（万）" in l0:
            df[col] = pd.to_numeric(df[col], errors="coerce").mul(10000).round(0).astype("Int64").astype(str).replace("<NA>", "")

    # 过滤第2列或第3列为空的行
    mask = (df.iloc[:, 1] == "") | (df.iloc[:, 2] == "")
    if mask.any():
        print("删除第2列或第3列为空的行：")
        df = df[~mask]

    df2 = df.copy()
    # 插入到第0列
    df2.insert(0, "id", f"{item.sheet_index}_" + (df2.index + 1000).astype(str))

    df2.to_csv(
        f"{current_dir}/assets/data/social_midia/{item.key}_{item.sheet_index}.txt",
        sep="\t",
        index=False,
        header=False,
        encoding="utf-8"
    )


def process_xiaohongshu_data(item: ExcelModel):
    """
        处理小红书数据，将excel中数据转成hive数仓中的观众数据表
        1. 不要表头
        2. 每列之间用制表符隔开
        3. NaN值或其它特殊符号用空字符串替换
        4. 将处理后的数据写入文本文件

        Args:
            item (ExcelModel): 包含抖音数据的ExcelModel对象
    """
    df = item.df
    # 第一列为合并列时，统一处理成上一行的值
    df.iloc[:, 0] = df.iloc[:, 0].ffill()

    # 第一列中有换行符的，根据换行符分割，并取第一个元素作为第一列值
    df.iloc[:, 0] = df.iloc[:, 0].apply(lambda x: x.split('\n')[0] if isinstance(x, str) else x)

    # 其它 NaN 统一转为空字符串
    #df_audience = item.df.fillna("")

    for col in df.columns:
        l0, l1 = col
        # NaN 值统一转为空字符串
        df[col] = df[col].fillna("")
        # "/" 替换为空字符串
        df[col] = df[col].replace("/", "", regex=False)


    # 过滤第2列或第3列为空的行
    mask = (df.iloc[:, 1] == "") | (df.iloc[:, 2] == "")
    if mask.any():
        print("删除第2列或第3列为空的行：")
        #print(df[mask])
        df = df[~mask]

    # 处理列值为 "/" 的情况
    #obj_cols = df_audience.select_dtypes(include="object").columns
    #df_audience[obj_cols] = df_audience[obj_cols].replace("/", "", regex=False)

    # 新增id字段，并输出到第一列。id生成规则：sheet_index_ + (index + 1000)
    df2 = df.copy()
    # 插入到第0列
    df2.insert(0, "id", f"{item.sheet_index}_" + (df2.index + 1000).astype(str))

    df2.to_csv(
        f"{current_dir}/assets/data/social_midia/{item.key}_{item.sheet_index}.txt",
        sep="\t",
        index=False,
        header=False,
        encoding="utf-8"
    )


def excel_to_csv():
    """
    将excel中社媒直播数据转换为csv文件
    """
    matched_sheets = read_social_midia_excel("~/Downloads/10000002_20260123145802.xlsx")
    #test_process_data(matched_sheets[0])

    for item in matched_sheets:
        if item.key == "douyin":
            process_douyin_data(item)
        elif item.key == "wechat":
            process_wechat_data(item)
        elif item.key == "xiaohongshu":
            process_xiaohongshu_data(item)
        else:
            print(f"未处理的sheet_index: {item.sheet_index}")


def get_match_result(row: pd.Series) -> tuple[str, int]:
    """
    匹配赛事id，生成row_id与match_id的映射关系
    """
    # 取1、2、3、4、11列数据并遍历
    match_id = 0
    id = row[0]
    # 非直播数据
    if "回顾" in row[1] or "回放" in row[1] or "录播" in row[1] or '录像' in row[10]:
        league_name = row[2].replace("回放", "").replace("回顾", "").replace("录播", "").replace("录像", "").replace("首播", "")
        # 将row[2] 中的 8-7 正则替换为 vs
        query = f"在{row[3]} 回放的 {league_name.replace(r'\d-\d', r'VS')}，对应的赛事ID是多少？"
    elif "节目" in row[1] or "日报" in row[1] or "情报" in row[1] or "抽签仪式" in row[2] or "发布会" in row[2]:
        query = ""
    else:
        league_name = row[1].replace("其他", "")
        query = f"在{row[3]}直播的 {league_name} {row[2]}，对应的赛事ID是多少？"

    query = query.replace("vs", "VS").replace("RB", "")
    if query:
        matched_result = query_match_info(query)
        #matched_result = []
        match_id = matched_result[0].get("match_id", 0) if matched_result and len(matched_result) > 0 else 0
    else:
        match_id = 0

    print(f"query: {query} ----> id: {id}, match_id: {match_id}")
    return id, match_id

def match_match_info():
    """
    匹配赛事id，生成row_id与match_id的映射关系
    """
    data_files = ["xiaohongshu_110.txt"]
    for data_file in data_files:
        df = pd.read_csv(
            f"{current_dir}/assets/data/social_midia/{data_file}",
            sep="\t",
            header=None,
            dtype=str,
            encoding="utf-8"
        )
        name_parts = data_file.split("_")
        tag = name_parts[0]
        sheet_index = name_parts[1].replace(".txt", "")
        matched_ids = []

        # 遍历每一行数据，获取row_id和match_id
        for index, row in df.iterrows():
            id, match_id = get_match_result(row)
            line = f"{tag}\t{id}\t{match_id}"
            matched_ids.append(line)

        # 将matched_ids写入文本文件
        with open(
              f"{current_dir}/assets/data/social_midia/{tag}_matched_id_{sheet_index}.txt",
              "w",
              encoding="utf-8"
          ) as f:
            f.write("\n".join(matched_ids))


def rematch_match_id():
    """
    重新匹配赛事id，生成row_id与match_id的映射关系

    """
    data_files = ["matched_id_103.txt"]
    # 读取文本文件
    for data_file in data_files:
        sheet_index = data_file.split("_")[2].replace(".txt", "")

        origin_df = pd.read_csv(
            f"{current_dir}/assets/data/social_midia/douyin_{sheet_index}.txt",
            sep="\t",
            header=None,
            dtype=str,
            encoding="utf-8"
        )

        df = pd.read_csv(
            f"{current_dir}/assets/data/social_midia/{data_file}",
            sep="\t",
            header=None,
            dtype=str,
            encoding="utf-8"
        )

        for index, row in df.iterrows():
            print("-"*50)
            #row_num = row[1]
            match_id = row[2]
            # 获取第2列的值，判断是否等于0，如果等于0，重新匹配
            if match_id == "0":
                print(f"第{index}行，match_id为0，需要重新匹配")
                # 获取origin_df中id等于df中第1列的值的行
                origin_row = origin_df[origin_df[0] == row[1]]
                if origin_row.empty:
                    print(f"第{index}行，id为{row[1]}，在origin_df中不存在")
                    continue

                id, match_id = get_match_result(origin_row.iloc[0])
                print(f"第{index}行，id为{id}，match_id重新匹配结果为{match_id}")
                df.iloc[index, 2] = match_id

        # 将df写入文本文件
        df.to_csv(
            f"{current_dir}/assets/data/social_midia/{data_file}",
            sep="\t",
            index=False,
            header=False,
            encoding="utf-8"
        )


def rematch_match_with_index(tag: str, id: str):
    """
    重新匹配赛事id，生成row_id与match_id的映射关系

    """
    id_parts = id.split("_")
    sheet_index = id_parts[0]
    row_num = id_parts[1]
    # 读取文本文件
    df = pd.read_csv(
        f"{current_dir}/assets/data/social_midia/{tag}_{sheet_index}.txt",
        sep="\t",
        header=None,
        dtype=str,
        encoding="utf-8"
    )

    # 获取指定行数据
    for index, row in df.iterrows():
        if row[0] == id:
            id, match_id = get_match_result(row)
            print(f"第{index}行，id为{id}，match_id重新匹配结果为{match_id}")
            break


if __name__ == '__main__':
    #excel_to_csv()
    print("="*50)

    # 匹配赛事id
    match_match_info()
    # 重新匹配赛事id
    #rematch_match_id()
    #rematch_match_with_index("douyin", "114_1305")
